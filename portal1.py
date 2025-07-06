from flask import Flask, request, session, redirect, url_for, render_template_string
from werkzeug.security import check_password_hash, generate_password_hash
import os
import re
from datetime import datetime
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
app.secret_key = os.urandom(24)

app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    SESSION_COOKIE_SECURE=False
)

EXCEL_PATH = "f:\\customers.xlsx"
USERNAME = "csc"
HASHED_PASSWORD = generate_password_hash("29375")

HTML_TEMPLATE = '''
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Portal</title>
<style>
    body {{
        font-family: 'Segoe UI', sans-serif;
        background: linear-gradient(135deg, #74ebd5, #ACB6E5);
        animation: fadeIn 1.5s ease-in;
        padding: 20px 0;
    }}
    .container {{
        background: rgba(255,255,255,0.9);
        padding: 40px;
        border-radius: 20px;
        box-shadow: 0 12px 30px rgba(0,0,0,0.2);
        max-width: 900px;
        margin: auto;
    }}
    h2 {{
        margin-top: 0;
        margin-bottom: 20px;
        border-bottom: 2px solid #74ebd5;
        padding-bottom: 8px;
    }}
    input[type="text"], input[type="password"], input[type="number"], input[type="date"] {{
        width: 90%;
        padding: 12px;
        margin: 12px 0;
        border: 1px solid #ccc;
        border-radius: 10px;
    }}
    input[type="submit"], button {{
        background: #74ebd5;
        color: white;
        border: none;
        padding: 14px 24px;
        border-radius: 10px;
        cursor: pointer;
        transition: background 0.4s;
        margin-top: 10px;
    }}
    input[type="submit"]:hover, button:hover {{
        background: #ACB6E5;
    }}
    table {{
        width: 100%;
        border-collapse: collapse;
        margin-top: 30px;
    }}
    th, td {{
        border: 1px solid #ccc;
        padding: 8px;
    }}
    th {{
        background: #74ebd5;
        color: white;
    }}
    a {{
        margin-left: 10px;
        font-size: 0.9em;
    }}
</style>
</head>
<body>
<div class="container">
    {content}
</div>
</body>
</html>
'''

def set_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    return response

@app.after_request
def after_request(response):
    return set_security_headers(response)

def get_excel_path_for_date(date_str):
    return f"f:\\regular_customers_{date_str}.xlsx"

# ========= ROOT REDIRECT =========
@app.route("/")
def index():
    return redirect(url_for("login"))

# ========= LOGIN / LOGOUT =========
@app.route("/login", methods=["GET", "POST"])
def login():
    message = "<h2>Login</h2>"
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        if username == USERNAME and check_password_hash(HASHED_PASSWORD, password):
            session["user"] = username
            return redirect(url_for("dashboard"))
        else:
            message += "<h3 style='color:red;'>Invalid credentials!</h3>"
    login_form = '''
        <form method="post">
            <input type="text" name="username" placeholder="Username" required><br>
            <input type="password" name="password" placeholder="Password" required><br>
            <input type="submit" value="Login">
        </form>
    '''
    return render_template_string(HTML_TEMPLATE.format(content=message + login_form))

@app.route("/logout")
def logout():
    session.pop("user", None)
    session.pop("date", None)
    return redirect(url_for("login"))

# ========= DASHBOARD =========
@app.route("/dashboard")
def dashboard():
    if "user" not in session:
        return redirect(url_for("login"))
    return render_template_string(HTML_TEMPLATE.format(content="""
        <h2>Welcome to Portal</h2>
        <a href="/register">Customer Registration</a><br><br>
        <a href="/regular">Regular Customer Work</a><br><br>
        <a href="/logout">Logout</a>
    """))

# ========= CUSTOMER REGISTRATION =========
@app.route("/register", methods=["GET", "POST"])
def register():
    if "user" not in session:
        return redirect(url_for("login"))
    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Phone Number", "Address"])
        wb.save(EXCEL_PATH)

    message = "<h2>Customer Registration</h2>"
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    if request.method == "POST":
        name = re.sub(r"[^\w\s\.-]", "", request.form.get("name", "").strip())
        phone = request.form.get("phone", "").strip() or None
        address = re.sub(r"[^\w\s\.,-]", "", request.form.get("address", "").strip()) or None
        ws.append([name, phone, address])
        wb.save(EXCEL_PATH)
        return redirect(url_for("register"))

    table_html = ""
    if request.args.get("show") == "1":
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) > 1:
            table_html = "<table><tr>" + "".join(f"<th>{col}</th>" for col in rows[0]) + "</tr>"
            for row in rows[1:]:
                table_html += "<tr>" + "".join(f"<td>{cell}</td>" for cell in row) + "</tr>"
            table_html += "</table>"
            table_html += '<form method="get"><button name="close" value="1">Close Table</button></form>'

    form_html = '''
        <form method="post">
            <input type="text" name="name" placeholder="Name" required><br>
            <input type="text" name="phone" placeholder="Phone Number"><br>
            <input type="text" name="address" placeholder="Address"><br>
            <input type="submit" value="Register">
        </form>
        <form method="get">
            <button name="show" value="1">Show Details</button>
        </form>
        <a href="/dashboard">Back to Dashboard</a> | <a href="/logout">Logout</a>
    '''
    return render_template_string(HTML_TEMPLATE.format(content=message + form_html + table_html))

# ========= REGULAR DAILY PORTAL =========
@app.route("/regular", methods=["GET", "POST"])
def regular():
    if "user" not in session:
        return redirect(url_for("login"))

    if "date" not in session:
        if request.method == "POST":
            selected_date = request.form.get("date")
            if selected_date:
                session["date"] = selected_date
                excel_path = get_excel_path_for_date(selected_date)
                if not os.path.exists(excel_path):
                    wb = Workbook()
                    ws = wb.active
                    ws.append(["Name", "Phone Number", "Address", "Amount GPay", "Amount Cash", "Work"])
                    wb.save(excel_path)
                return redirect(url_for("regular"))
        date_form = '''
            <h2>Select Date for Work</h2>
            <form method="post">
                <input type="date" name="date" required><br>
                <input type="submit" value="Start Day">
            </form>
            <a href="/logout">Logout</a>
        '''
        return render_template_string(HTML_TEMPLATE.format(content=date_form))

    selected_date = session["date"]
    excel_path = get_excel_path_for_date(selected_date)

    edit_row = request.args.get("edit")
    close_table = request.args.get("close")
    message = f"<h2>Regular Customer Portal - {selected_date}</h2>"
    form_values = {"name": "", "phone": "", "address": "", "amount_gpay": "", "amount_cash": "", "work": ""}

    wb = load_workbook(excel_path)
    ws = wb.active

    if edit_row is not None and edit_row.isdigit():
        idx = int(edit_row) + 2
        if idx <= ws.max_row:
            row_data = [str(ws.cell(row=idx, column=col).value or '') for col in range(1,7)]
            form_values = dict(zip(["name", "phone", "address", "amount_gpay", "amount_cash", "work"], row_data))

    if request.method == "POST" and "date" not in request.form:
        name = re.sub(r"[^\w\s\.-]", "", request.form.get("name", "").strip())
        phone = request.form.get("phone", "").strip() or None
        address = re.sub(r"[^\w\s\.,-]", "", request.form.get("address", "").strip()) or None
        work = re.sub(r"[^\w\s\.,-]", "", request.form.get("work", "").strip()) or None
        try:
            amount_gpay = float(request.form.get("amount_gpay", "0").strip() or 0)
            amount_cash = float(request.form.get("amount_cash", "0").strip() or 0)
        except ValueError:
            amount_gpay, amount_cash = 0, 0

        if edit_row is not None and edit_row.isdigit():
            idx = int(edit_row) + 2
            ws.cell(row=idx, column=1, value=name)
            ws.cell(row=idx, column=2, value=phone)
            ws.cell(row=idx, column=3, value=address)
            ws.cell(row=idx, column=4, value=amount_gpay)
            ws.cell(row=idx, column=5, value=amount_cash)
            ws.cell(row=idx, column=6, value=work)
        else:
            ws.append([name, phone, address, amount_gpay, amount_cash, work])
        wb.save(excel_path)
        return redirect(url_for("regular"))

    total_gpay, total_cash = 0, 0
    table_html = ""
    if request.args.get("show") == "1" and not close_table:
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) > 1:
            table_html = "<table><tr>" + "".join(f"<th>{col}</th>" for col in rows[0]) + "<th>Action</th></tr>"
            for idx, row in enumerate(rows[1:]):
                total_gpay += float(row[3] or 0)
                total_cash += float(row[4] or 0)
                action = f'<a href="/regular?edit={idx}">Edit</a>'
                table_html += "<tr>" + "".join(f"<td>{cell}</td>" for cell in row) + f"<td>{action}</td></tr>"
            table_html += f"</table><h3>Total GPay: {total_gpay} | Total Cash: {total_cash}</h3>"
            table_html += '<form method="get"><button name="close" value="1">Close Table</button></form>'

    form_html = '''
        <form method="post">
            <input type="text" name="name" placeholder="Name" value="{name}" required><br>
            <input type="text" name="phone" placeholder="Phone Number" value="{phone}"><br>
            <input type="text" name="address" placeholder="Address" value="{address}"><br>
            <input type="number" step="0.01" name="amount_gpay" placeholder="Amount GPay" value="{amount_gpay}"><br>
            <input type="number" step="0.01" name="amount_cash" placeholder="Amount Cash" value="{amount_cash}"><br>
            <input type="text" name="work" placeholder="Work Description" value="{work}"><br>
            <input type="submit" value="{button}">
        </form>
        <form method="get">
            <button name="show" value="1">Show Details</button>
        </form>
        <a href="/dashboard">Back to Dashboard</a> | <a href="/logout">Logout</a>
    '''.format(**form_values, button=("Update" if edit_row else "Submit"))

    return render_template_string(HTML_TEMPLATE.format(content=message + form_html + table_html))

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=False)=
